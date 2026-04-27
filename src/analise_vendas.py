from pathlib import Path

import matplotlib.pyplot as plt
import pandas as pd
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parents[1]
PASTA_ENTRADA = BASE_DIR / "entrada"
PASTA_SAIDA = BASE_DIR / "saida"
PASTA_DADOS_EXEMPLO = BASE_DIR / "dados"

CAMINHO_EXCEL = PASTA_SAIDA / "relatorio_vendas.xlsx"
CAMINHO_GRAFICO_PRODUTO = PASTA_SAIDA / "grafico_faturamento_produto.png"
CAMINHO_GRAFICO_MES = PASTA_SAIDA / "grafico_vendas_mes.png"
CAMINHO_GRAFICO_CLIENTES = PASTA_SAIDA / "grafico_top_clientes.png"
CAMINHO_GRAFICO_PAGAMENTO = PASTA_SAIDA / "grafico_formas_pagamento.png"
CAMINHO_DASHBOARD = PASTA_SAIDA / "dashboard.html"

COLUNAS_OBRIGATORIAS = {
    "data_venda",
    "produto",
    "categoria",
    "quantidade",
    "preco_unitario",
    "cliente",
    "cidade",
    "estado",
    "forma_pagamento",
}


class ErroArquivoEntrada(Exception):
    pass


def formatar_moeda(valor: float) -> str:
    return f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def formatar_moeda_curta(valor: float) -> str:
    if valor >= 1000:
        return f"R$ {valor / 1000:.1f} mil".replace(".", ",")
    return formatar_moeda(valor)


def localizar_arquivo_entrada() -> Path:
    arquivos = sorted(
        [
            *PASTA_ENTRADA.glob("*.csv"),
            *PASTA_ENTRADA.glob("*.xlsx"),
            *PASTA_ENTRADA.glob("*.xls"),
        ]
    )
    if not arquivos:
        exemplo = PASTA_DADOS_EXEMPLO / "vendas.csv"
        raise ErroArquivoEntrada(
            "Nenhuma planilha foi encontrada na pasta 'entrada'.\n"
            f"Coloque um arquivo CSV ou Excel nessa pasta. Exemplo: {exemplo}"
        )
    return arquivos[0]


def ler_planilha(caminho_arquivo: Path) -> pd.DataFrame:
    if caminho_arquivo.suffix.lower() == ".csv":
        return pd.read_csv(caminho_arquivo)
    return pd.read_excel(caminho_arquivo)


def validar_colunas(df: pd.DataFrame) -> None:
    colunas_encontradas = set(df.columns)
    colunas_faltantes = sorted(COLUNAS_OBRIGATORIAS - colunas_encontradas)
    if colunas_faltantes:
        raise ErroArquivoEntrada(
            "A planilha esta sem colunas obrigatorias.\n"
            f"Colunas faltantes: {', '.join(colunas_faltantes)}\n"
            f"Colunas esperadas: {', '.join(sorted(COLUNAS_OBRIGATORIAS))}"
        )


def carregar_e_tratar_dados(caminho_arquivo: Path) -> pd.DataFrame:
    df = ler_planilha(caminho_arquivo)
    df.columns = [str(col).strip().lower() for col in df.columns]
    validar_colunas(df)

    linhas_originais = len(df)
    df["data_venda"] = pd.to_datetime(df["data_venda"], errors="coerce")
    df["quantidade"] = pd.to_numeric(df["quantidade"], errors="coerce")
    df["preco_unitario"] = pd.to_numeric(df["preco_unitario"], errors="coerce")

    df = df.dropna(subset=["data_venda", "produto", "quantidade", "preco_unitario"])
    df = df[df["quantidade"] > 0]
    df = df[df["preco_unitario"] > 0]

    if df.empty:
        raise ErroArquivoEntrada(
            "A planilha foi lida, mas nenhuma venda valida foi encontrada."
        )

    colunas_texto = ["produto", "categoria", "cliente", "cidade", "estado", "forma_pagamento"]
    for coluna in colunas_texto:
        df[coluna] = df[coluna].astype(str).str.strip()

    df["faturamento"] = df["quantidade"] * df["preco_unitario"]
    df["mes"] = df["data_venda"].dt.to_period("M").astype(str)
    df["ano"] = df["data_venda"].dt.year
    df["linhas_removidas_tratamento"] = linhas_originais - len(df)

    return df.sort_values("data_venda").reset_index(drop=True)


def calcular_metricas(df: pd.DataFrame) -> dict:
    vendas_por_produto = (
        df.groupby("produto", as_index=False)
        .agg(quantidade=("quantidade", "sum"), faturamento=("faturamento", "sum"))
        .sort_values("faturamento", ascending=False)
    )
    vendas_por_mes = (
        df.groupby("mes", as_index=False)
        .agg(quantidade=("quantidade", "sum"), faturamento=("faturamento", "sum"))
        .sort_values("mes")
    )
    vendas_por_cliente = (
        df.groupby("cliente", as_index=False)
        .agg(quantidade=("quantidade", "sum"), faturamento=("faturamento", "sum"))
        .sort_values("faturamento", ascending=False)
    )
    vendas_por_estado = (
        df.groupby("estado", as_index=False)
        .agg(quantidade=("quantidade", "sum"), faturamento=("faturamento", "sum"))
        .sort_values("faturamento", ascending=False)
    )
    vendas_por_pagamento = (
        df.groupby("forma_pagamento", as_index=False)
        .agg(quantidade=("quantidade", "sum"), faturamento=("faturamento", "sum"))
        .sort_values("faturamento", ascending=False)
    )

    faturamento_total = df["faturamento"].sum()
    ticket_medio = df["faturamento"].mean()
    quantidade_total = int(df["quantidade"].sum())
    pedidos_total = len(df)
    produto_mais_vendido = vendas_por_produto.sort_values(
        "quantidade", ascending=False
    ).iloc[0]["produto"]
    melhor_mes = vendas_por_mes.sort_values("faturamento", ascending=False).iloc[0]["mes"]
    melhor_cliente = vendas_por_cliente.iloc[0]["cliente"]

    vendas_por_mes["crescimento_percentual"] = (
        vendas_por_mes["faturamento"].pct_change().fillna(0) * 100
    )

    indicadores = pd.DataFrame(
        [
            ["Faturamento total", faturamento_total],
            ["Ticket medio", ticket_medio],
            ["Quantidade total vendida", quantidade_total],
            ["Total de pedidos", pedidos_total],
            ["Produto mais vendido", produto_mais_vendido],
            ["Melhor mes de venda", melhor_mes],
            ["Cliente com maior faturamento", melhor_cliente],
        ],
        columns=["indicador", "valor"],
    )

    return {
        "indicadores": indicadores,
        "vendas_por_produto": vendas_por_produto,
        "vendas_por_mes": vendas_por_mes,
        "vendas_por_cliente": vendas_por_cliente,
        "vendas_por_estado": vendas_por_estado,
        "vendas_por_pagamento": vendas_por_pagamento,
    }


def salvar_graficos(metricas: dict) -> None:
    plt.style.use("seaborn-v0_8-whitegrid")
    cores = ["#1f4e78", "#0f766e", "#d97706", "#b91c1c", "#475569", "#7c3aed"]

    produto = metricas["vendas_por_produto"].head(8).sort_values("faturamento")
    fig, ax = plt.subplots(figsize=(12, 6.5), facecolor="white")
    barras = ax.barh(produto["produto"], produto["faturamento"], color="#1f4e78")
    ax.set_title("Faturamento por Produto", fontsize=18, weight="bold", pad=18, loc="left")
    ax.text(
        0,
        1.02,
        "Produtos ordenados por receita gerada",
        transform=ax.transAxes,
        color="#64748b",
        fontsize=11,
    )
    ax.set_xlabel("Faturamento")
    ax.xaxis.set_major_formatter(lambda valor, _: f"R$ {valor / 1000:.0f} mil")
    ax.bar_label(
        barras,
        labels=[formatar_moeda_curta(valor) for valor in produto["faturamento"]],
        padding=6,
        fontsize=10,
        color="#172033",
    )
    ax.grid(axis="x", alpha=0.22)
    ax.grid(axis="y", visible=False)
    ax.spines[["top", "right", "left", "bottom"]].set_visible(False)
    ax.tick_params(axis="y", length=0)
    fig.tight_layout()
    fig.savefig(CAMINHO_GRAFICO_PRODUTO, dpi=170)
    plt.close(fig)

    mes = metricas["vendas_por_mes"]
    fig, ax = plt.subplots(figsize=(12, 6.5), facecolor="white")
    ax.plot(mes["mes"], mes["faturamento"], marker="o", linewidth=2.6, color="#0f766e")
    ax.fill_between(mes["mes"], mes["faturamento"], alpha=0.12, color="#0f766e")
    ax.set_title("Evolucao do Faturamento Mensal", fontsize=18, weight="bold", pad=18, loc="left")
    ax.text(
        0,
        1.02,
        "Tendencia mensal de vendas",
        transform=ax.transAxes,
        color="#64748b",
        fontsize=11,
    )
    ax.set_xlabel("Mes")
    ax.set_ylabel("Faturamento")
    ax.tick_params(axis="x", rotation=45)
    ax.yaxis.set_major_formatter(lambda valor, _: f"R$ {valor / 1000:.0f} mil")
    for _, linha in mes.iterrows():
        ax.annotate(
            formatar_moeda_curta(linha["faturamento"]),
            (linha["mes"], linha["faturamento"]),
            textcoords="offset points",
            xytext=(0, 9),
            ha="center",
            fontsize=9,
            color="#172033",
        )
    ax.grid(axis="y", alpha=0.22)
    ax.grid(axis="x", visible=False)
    ax.spines[["top", "right", "left", "bottom"]].set_visible(False)
    fig.tight_layout()
    fig.savefig(CAMINHO_GRAFICO_MES, dpi=170)
    plt.close(fig)

    clientes = metricas["vendas_por_cliente"].head(5).sort_values("faturamento")
    fig, ax = plt.subplots(figsize=(12, 6.5), facecolor="white")
    barras = ax.barh(clientes["cliente"], clientes["faturamento"], color="#b91c1c")
    ax.set_title("Top 5 Clientes por Faturamento", fontsize=18, weight="bold", pad=18, loc="left")
    ax.text(
        0,
        1.02,
        "Clientes com maior impacto no resultado",
        transform=ax.transAxes,
        color="#64748b",
        fontsize=11,
    )
    ax.set_xlabel("Faturamento")
    ax.xaxis.set_major_formatter(lambda valor, _: f"R$ {valor / 1000:.0f} mil")
    ax.bar_label(
        barras,
        labels=[formatar_moeda_curta(valor) for valor in clientes["faturamento"]],
        padding=6,
        fontsize=10,
        color="#172033",
    )
    ax.grid(axis="x", alpha=0.22)
    ax.grid(axis="y", visible=False)
    ax.spines[["top", "right", "left", "bottom"]].set_visible(False)
    ax.tick_params(axis="y", length=0)
    fig.tight_layout()
    fig.savefig(CAMINHO_GRAFICO_CLIENTES, dpi=170)
    plt.close(fig)

    pagamento = metricas["vendas_por_pagamento"]
    fig, ax = plt.subplots(figsize=(9, 6.5), facecolor="white")
    wedges, texts, autotexts = ax.pie(
        pagamento["faturamento"],
        labels=pagamento["forma_pagamento"],
        autopct="%1.1f%%",
        startangle=90,
        colors=cores[: len(pagamento)],
        wedgeprops={"width": 0.42, "edgecolor": "white", "linewidth": 2},
        textprops={"fontsize": 10, "color": "#172033"},
    )
    for texto in autotexts:
        texto.set_color("white")
        texto.set_weight("bold")
    ax.set_title("Faturamento por Forma de Pagamento", fontsize=18, weight="bold", pad=18)
    ax.text(0, 0, "Pagamentos", ha="center", va="center", fontsize=12, weight="bold", color="#475569")
    fig.tight_layout()
    fig.savefig(CAMINHO_GRAFICO_PAGAMENTO, dpi=170)
    plt.close(fig)


def moeda_excel(ws, colunas: list[str]) -> None:
    cabecalhos = {cell.value: cell.column for cell in ws[1]}
    for coluna in colunas:
        indice = cabecalhos.get(coluna)
        if indice:
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=indice).number_format = '"R$" #,##0.00'


def criar_card_excel(ws, intervalo: str, titulo: str, valor: str, cor: str) -> None:
    ws.merge_cells(intervalo)
    celula = ws[intervalo.split(":")[0]]
    celula.value = f"{titulo}\n{valor}"
    celula.fill = PatternFill("solid", fgColor=cor)
    celula.font = Font(color="FFFFFF", bold=True, size=14)
    celula.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    borda = Side(style="thin", color="FFFFFF")
    for row in ws[intervalo]:
        for cell in row:
            cell.border = Border(left=borda, right=borda, top=borda, bottom=borda)


def adicionar_imagem_excel(ws, caminho: Path, celula: str, largura: int) -> None:
    if not caminho.exists():
        return
    imagem = Image(str(caminho))
    proporcao = largura / imagem.width
    imagem.width = largura
    imagem.height = int(imagem.height * proporcao)
    ws.add_image(imagem, celula)


def criar_aba_dashboard_excel(workbook, metricas: dict) -> None:
    if "Dashboard" in workbook.sheetnames:
        del workbook["Dashboard"]

    ws = workbook.create_sheet("Dashboard", 0)
    ws.sheet_view.showGridLines = False

    for coluna in range(1, 16):
        ws.column_dimensions[get_column_letter(coluna)].width = 13
    for linha in range(1, 42):
        ws.row_dimensions[linha].height = 24

    indicadores = metricas["indicadores"].set_index("indicador")["valor"].to_dict()

    ws.merge_cells("B2:N3")
    titulo = ws["B2"]
    titulo.value = "Dashboard Executivo de Vendas"
    titulo.font = Font(size=22, bold=True, color="1F4E78")
    titulo.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("B4:N4")
    subtitulo = ws["B4"]
    subtitulo.value = "Indicadores, rankings e graficos gerados automaticamente a partir da planilha de entrada."
    subtitulo.font = Font(size=11, color="64748B")
    subtitulo.alignment = Alignment(horizontal="left", vertical="center")

    criar_card_excel(
        ws,
        "B6:D8",
        "Faturamento total",
        formatar_moeda(indicadores["Faturamento total"]),
        "1F4E78",
    )
    criar_card_excel(
        ws,
        "E6:G8",
        "Ticket medio",
        formatar_moeda(indicadores["Ticket medio"]),
        "0F766E",
    )
    criar_card_excel(
        ws,
        "H6:J8",
        "Qtd. vendida",
        str(int(indicadores["Quantidade total vendida"])),
        "D97706",
    )
    criar_card_excel(
        ws,
        "K6:M8",
        "Total pedidos",
        str(int(indicadores["Total de pedidos"])),
        "B91C1C",
    )

    ws.merge_cells("B10:G10")
    ws["B10"].value = "Graficos principais"
    ws["B10"].font = Font(size=14, bold=True, color="172033")

    adicionar_imagem_excel(ws, CAMINHO_GRAFICO_MES, "B12", 585)
    adicionar_imagem_excel(ws, CAMINHO_GRAFICO_PRODUTO, "I12", 585)
    adicionar_imagem_excel(ws, CAMINHO_GRAFICO_CLIENTES, "B30", 585)
    adicionar_imagem_excel(ws, CAMINHO_GRAFICO_PAGAMENTO, "I30", 430)


def criar_relatorio_excel(df: pd.DataFrame, metricas: dict) -> None:
    with pd.ExcelWriter(CAMINHO_EXCEL, engine="openpyxl") as writer:
        metricas["indicadores"].to_excel(writer, index=False, sheet_name="Resumo Executivo")
        metricas["vendas_por_produto"].to_excel(
            writer, index=False, sheet_name="Produtos"
        )
        metricas["vendas_por_mes"].to_excel(writer, index=False, sheet_name="Meses")
        metricas["vendas_por_cliente"].to_excel(writer, index=False, sheet_name="Clientes")
        metricas["vendas_por_estado"].to_excel(writer, index=False, sheet_name="Estados")
        metricas["vendas_por_pagamento"].to_excel(
            writer, index=False, sheet_name="Pagamentos"
        )
        df.drop(columns=["linhas_removidas_tratamento"]).to_excel(
            writer, index=False, sheet_name="Base Tratada"
        )

        workbook = writer.book
        borda = Side(style="thin", color="D9E2EC")
        for worksheet in workbook.worksheets:
            worksheet.freeze_panes = "A2"
            worksheet.sheet_view.showGridLines = False
            for cell in worksheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1F4E78")
                cell.alignment = Alignment(horizontal="center")
                cell.border = Border(bottom=borda)

            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.border = Border(bottom=borda)
                    cell.alignment = Alignment(vertical="center")

            for column_cells in worksheet.columns:
                max_length = max(len(str(cell.value or "")) for cell in column_cells)
                width = min(max(max_length + 3, 14), 42)
                worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width

            moeda_excel(worksheet, ["faturamento", "preco_unitario"])

        resumo_ws = workbook["Resumo Executivo"]
        resumo_ws.column_dimensions["A"].width = 32
        resumo_ws.column_dimensions["B"].width = 28
        criar_aba_dashboard_excel(workbook, metricas)

        produto_ws = workbook["Produtos"]
        produto_chart = BarChart()
        produto_chart.title = "Faturamento por Produto"
        produto_chart.y_axis.title = "Faturamento"
        produto_chart.x_axis.title = "Produto"
        produto_chart.add_data(
            Reference(produto_ws, min_col=3, min_row=1, max_row=produto_ws.max_row),
            titles_from_data=True,
        )
        produto_chart.set_categories(
            Reference(produto_ws, min_col=1, min_row=2, max_row=produto_ws.max_row)
        )
        produto_chart.height = 8
        produto_chart.width = 18
        produto_ws.add_chart(produto_chart, "E2")

        mes_ws = workbook["Meses"]
        mes_chart = LineChart()
        mes_chart.title = "Faturamento por Mes"
        mes_chart.y_axis.title = "Faturamento"
        mes_chart.x_axis.title = "Mes"
        mes_chart.add_data(
            Reference(mes_ws, min_col=3, min_row=1, max_row=mes_ws.max_row),
            titles_from_data=True,
        )
        mes_chart.set_categories(
            Reference(mes_ws, min_col=1, min_row=2, max_row=mes_ws.max_row)
        )
        mes_chart.height = 8
        mes_chart.width = 18
        mes_ws.add_chart(mes_chart, "F2")


def tabela_html(df: pd.DataFrame, colunas_moeda: list[str] | None = None) -> str:
    colunas_moeda = colunas_moeda or []
    formatters = {col: formatar_moeda for col in colunas_moeda}
    return df.to_html(index=False, classes="tabela", formatters=formatters)


def criar_dashboard_html(metricas: dict, arquivo_entrada: Path) -> None:
    indicadores = metricas["indicadores"].set_index("indicador")["valor"].to_dict()
    produto_html = tabela_html(
        metricas["vendas_por_produto"].head(8), colunas_moeda=["faturamento"]
    )
    cliente_html = tabela_html(
        metricas["vendas_por_cliente"].head(8), colunas_moeda=["faturamento"]
    )
    estado_html = tabela_html(metricas["vendas_por_estado"], colunas_moeda=["faturamento"])
    pagamento_html = tabela_html(
        metricas["vendas_por_pagamento"], colunas_moeda=["faturamento"]
    )

    html = f"""<!doctype html>
<html lang="pt-BR">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Dashboard de Vendas</title>
  <style>
    :root {{
      --azul: #1f4e78;
      --verde: #0f766e;
      --vermelho: #b91c1c;
      --texto: #172033;
      --muted: #64748b;
      --linha: #d9e2ec;
      --fundo: #f5f7fb;
    }}
    * {{ box-sizing: border-box; }}
    body {{ margin: 0; font-family: Arial, sans-serif; background: var(--fundo); color: var(--texto); }}
    header {{ background: #ffffff; border-bottom: 1px solid var(--linha); }}
    .topo {{ max-width: 1220px; margin: 0 auto; padding: 24px 20px; display: flex; justify-content: space-between; gap: 20px; align-items: center; }}
    .marca {{ font-weight: 700; color: var(--azul); font-size: 15px; text-transform: uppercase; letter-spacing: .08em; }}
    h1 {{ margin: 8px 0 4px; font-size: 32px; line-height: 1.15; }}
    .subtitulo {{ margin: 0; color: var(--muted); }}
    .fonte {{ color: var(--muted); font-size: 13px; text-align: right; }}
    main {{ max-width: 1220px; margin: 0 auto; padding: 26px 20px 42px; }}
    .cards {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(190px, 1fr)); gap: 14px; margin-bottom: 22px; }}
    .card {{ background: #fff; border: 1px solid var(--linha); border-radius: 8px; padding: 18px; box-shadow: 0 8px 22px rgba(15, 23, 42, .06); min-height: 100px; }}
    .card span {{ display: block; color: var(--muted); font-size: 13px; margin-bottom: 10px; }}
    .card strong {{ display: block; font-size: 21px; line-height: 1.18; }}
    .grid {{ display: grid; grid-template-columns: repeat(2, minmax(0, 1fr)); gap: 18px; }}
    section {{ background: #fff; border: 1px solid var(--linha); border-radius: 8px; padding: 18px; overflow: auto; }}
    section h2 {{ margin: 0 0 14px; font-size: 18px; }}
    img {{ width: 100%; height: auto; display: block; }}
    .tabela {{ width: 100%; border-collapse: collapse; font-size: 14px; }}
    .tabela th, .tabela td {{ border-bottom: 1px solid #e7edf5; padding: 10px; text-align: left; white-space: nowrap; }}
    .tabela th {{ background: var(--azul); color: #fff; }}
    @media (max-width: 800px) {{
      .topo {{ display: block; }}
      .fonte {{ text-align: left; margin-top: 12px; }}
      .grid {{ grid-template-columns: 1fr; }}
    }}
  </style>
</head>
<body>
  <header>
    <div class="topo">
      <div>
        <div class="marca">Sales Report Automation</div>
        <h1>Dashboard de Vendas</h1>
        <p class="subtitulo">Indicadores comerciais gerados automaticamente.</p>
      </div>
      <div class="fonte">Arquivo analisado<br><strong>{arquivo_entrada.name}</strong></div>
    </div>
  </header>
  <main>
    <div class="cards">
      <div class="card"><span>Faturamento total</span><strong>{formatar_moeda(indicadores["Faturamento total"])}</strong></div>
      <div class="card"><span>Ticket medio</span><strong>{formatar_moeda(indicadores["Ticket medio"])}</strong></div>
      <div class="card"><span>Quantidade vendida</span><strong>{int(indicadores["Quantidade total vendida"])}</strong></div>
      <div class="card"><span>Total de pedidos</span><strong>{int(indicadores["Total de pedidos"])}</strong></div>
      <div class="card"><span>Produto mais vendido</span><strong>{indicadores["Produto mais vendido"]}</strong></div>
      <div class="card"><span>Melhor mes</span><strong>{indicadores["Melhor mes de venda"]}</strong></div>
    </div>
    <div class="grid">
      <section><h2>Faturamento por produto</h2><img src="grafico_faturamento_produto.png" alt="Faturamento por produto"></section>
      <section><h2>Evolucao mensal</h2><img src="grafico_vendas_mes.png" alt="Vendas por mes"></section>
      <section><h2>Top clientes</h2><img src="grafico_top_clientes.png" alt="Top clientes"></section>
      <section><h2>Formas de pagamento</h2><img src="grafico_formas_pagamento.png" alt="Formas de pagamento"></section>
      <section><h2>Ranking de produtos</h2>{produto_html}</section>
      <section><h2>Ranking de clientes</h2>{cliente_html}</section>
      <section><h2>Vendas por estado</h2>{estado_html}</section>
      <section><h2>Formas de pagamento</h2>{pagamento_html}</section>
    </div>
  </main>
</body>
</html>"""
    CAMINHO_DASHBOARD.write_text(html, encoding="utf-8")


def main() -> None:
    PASTA_ENTRADA.mkdir(exist_ok=True)
    PASTA_SAIDA.mkdir(exist_ok=True)

    try:
        arquivo_entrada = localizar_arquivo_entrada()
        df = carregar_e_tratar_dados(arquivo_entrada)
        metricas = calcular_metricas(df)
        salvar_graficos(metricas)
        criar_relatorio_excel(df, metricas)
        criar_dashboard_html(metricas, arquivo_entrada)
    except ErroArquivoEntrada as erro:
        print("\nERRO NA ENTRADA DE DADOS")
        print(str(erro))
        raise SystemExit(1) from erro

    indicadores = metricas["indicadores"]
    print("\nAnalise concluida com sucesso.")
    print(f"Arquivo analisado: {arquivo_entrada.name}")
    print(indicadores.to_string(index=False))
    print(f"\nArquivos gerados em: {PASTA_SAIDA}")
    print("- relatorio_vendas.xlsx")
    print("- grafico_faturamento_produto.png")
    print("- grafico_vendas_mes.png")
    print("- grafico_top_clientes.png")
    print("- grafico_formas_pagamento.png")
    print("- dashboard.html")


if __name__ == "__main__":
    main()
